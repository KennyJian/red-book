from flask import Flask, request, jsonify, send_from_directory, send_file
from flask_cors import CORS
import os
import json
import threading
import queue as thread_queue
import time
from datetime import datetime
# 使用 MediaCrawler 爬虫
try:
    from xhs_crawler_adapter import XHSCrawlerAdapter as XHSCrawler
    print("=" * 50)
    print("[成功] 使用 MediaCrawler 爬虫 (Playwright)")
    print("=" * 50)
except ImportError as e:
    print("=" * 50)
    print("[错误] 无法导入 MediaCrawler 爬虫，错误:", str(e))
    print("")
    print("解决方案：")
    print("1. 安装依赖: pip install -r requirements.txt")
    print("2. 确保 playwright 已安装: playwright install chromium")
    print("=" * 50)
    raise ImportError("MediaCrawler 爬虫依赖未安装，请先安装依赖") from e

app = Flask(__name__, static_folder='frontend', static_url_path='')
CORS(app)

# 数据存储目录
DATA_DIR = os.path.join(os.path.dirname(__file__), 'data')
USERS_DIR = os.path.join(DATA_DIR, 'users')
KEYWORDS_DIR = os.path.join(DATA_DIR, 'keywords')
COMMENTS_DIR = os.path.join(DATA_DIR, 'comments')

# 确保目录存在
os.makedirs(USERS_DIR, exist_ok=True)
os.makedirs(KEYWORDS_DIR, exist_ok=True)
os.makedirs(COMMENTS_DIR, exist_ok=True)

# 全局爬虫实例和状态
crawler_status = {
    'running': False,
    'current_keyword': '',
    'progress': 0,
    'total_users': 0,
    'message': ''
}

# Playwright 链接浏览器：用于在独立窗口中打开用户主页/帖子链接
_link_browser_ready = threading.Event()
_link_browser_thread = None
_link_browser_url_queue = thread_queue.Queue()

@app.route('/')
def index():
    return send_from_directory('frontend', 'index.html')

@app.route('/api/keywords', methods=['GET'])
def get_keywords():
    """获取所有关键词配置"""
    keywords_file = os.path.join(KEYWORDS_DIR, 'keywords.json')
    if os.path.exists(keywords_file):
        with open(keywords_file, 'r', encoding='utf-8') as f:
            keywords = json.load(f)
        return jsonify(keywords)
    return jsonify([])

@app.route('/api/keywords', methods=['POST'])
def add_keyword():
    """添加关键词"""
    data = request.json
    keyword = data.get('keyword', '').strip()
    
    if not keyword:
        return jsonify({'error': '关键词不能为空'}), 400
    
    keywords_file = os.path.join(KEYWORDS_DIR, 'keywords.json')
    keywords = []
    if os.path.exists(keywords_file):
        with open(keywords_file, 'r', encoding='utf-8') as f:
            keywords = json.load(f)
    
    if keyword not in keywords:
        keywords.append(keyword)
        with open(keywords_file, 'w', encoding='utf-8') as f:
            json.dump(keywords, f, ensure_ascii=False, indent=2)
    
    return jsonify({'success': True, 'keywords': keywords})

@app.route('/api/keywords/<keyword>', methods=['DELETE'])
def delete_keyword(keyword):
    """删除关键词"""
    keywords_file = os.path.join(KEYWORDS_DIR, 'keywords.json')
    if os.path.exists(keywords_file):
        with open(keywords_file, 'r', encoding='utf-8') as f:
            keywords = json.load(f)
        
        if keyword in keywords:
            keywords.remove(keyword)
            with open(keywords_file, 'w', encoding='utf-8') as f:
                json.dump(keywords, f, ensure_ascii=False, indent=2)
    
    return jsonify({'success': True})

@app.route('/api/crawl', methods=['POST'])
def start_crawl():
    """开始爬取"""
    global crawler_status
    
    if crawler_status['running']:
        return jsonify({'error': '爬虫正在运行中'}), 400
    
    data = request.json
    keywords = data.get('keywords', [])
    # 新增配置参数
    max_notes = data.get('max_notes', 100)  # 最多抓取的帖子数
    max_comments = data.get('max_comments', 100)  # 每个帖子最多抓取的评论数
    comment_filter_keywords = data.get('comment_filter_keywords', [])  # 评论过滤关键词
    
    if not keywords:
        return jsonify({'error': '请至少添加一个关键词'}), 400
    
    # 启动爬虫线程
    thread = threading.Thread(target=run_crawler, args=(keywords, max_notes, max_comments, comment_filter_keywords))
    thread.daemon = True
    thread.start()
    
    return jsonify({'success': True, 'message': '爬虫已启动'})

@app.route('/api/status', methods=['GET'])
def get_status():
    """获取爬虫状态"""
    return jsonify(crawler_status)

@app.route('/api/users', methods=['GET'])
def get_users():
    """获取所有用户数据"""
    users = []
    for filename in os.listdir(USERS_DIR):
        if filename.endswith('.json'):
            filepath = os.path.join(USERS_DIR, filename)
            with open(filepath, 'r', encoding='utf-8') as f:
                user_data = json.load(f)
                users.append(user_data)
    
    # 按时间排序
    users.sort(key=lambda x: x.get('crawl_time', ''), reverse=True)
    return jsonify(users)

@app.route('/api/users/stats', methods=['GET'])
def get_stats():
    """获取统计信息"""
    total_users = len([f for f in os.listdir(USERS_DIR) if f.endswith('.json')])
    return jsonify({
        'total_users': total_users,
        'data_dir': DATA_DIR
    })


def _load_users_list():
    """加载用户列表（与 get_users 逻辑一致）"""
    users = []
    for filename in os.listdir(USERS_DIR):
        if filename.endswith('.json'):
            filepath = os.path.join(USERS_DIR, filename)
            with open(filepath, 'r', encoding='utf-8') as f:
                users.append(json.load(f))
    users.sort(key=lambda x: x.get('crawl_time', ''), reverse=True)
    return users


@app.route('/api/users/export/excel', methods=['GET'])
def export_users_excel():
    """导出用户数据为 Excel 文件"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        return jsonify({'error': '请先安装 openpyxl: pip install openpyxl'}), 500

    users = _load_users_list()
    if not users:
        return jsonify({'error': '暂无用户数据可导出'}), 400

    wb = Workbook()
    # 工作表1：用户列表
    ws1 = wb.active
    ws1.title = "用户列表"
    headers1 = ["用户ID", "昵称", "主页简介", "主页链接", "来源关键词", "爬取时间", "评论数"]
    for col, h in enumerate(headers1, 1):
        ws1.cell(row=1, column=col, value=h)
        ws1.cell(row=1, column=col).font = Font(bold=True)
    for row, u in enumerate(users, 2):
        ws1.cell(row=row, column=1, value=u.get("user_id", ""))
        ws1.cell(row=row, column=2, value=u.get("nickname", ""))
        ws1.cell(row=row, column=3, value=u.get("desc", "") or u.get("user_desc", ""))
        ws1.cell(row=row, column=4, value=u.get("user_url", ""))
        ws1.cell(row=row, column=5, value=u.get("keyword", ""))
        ws1.cell(row=row, column=6, value=u.get("crawl_time", ""))
        comments = u.get("comments") or []
        ws1.cell(row=row, column=7, value=len(comments))
    for col in range(1, 8):
        ws1.column_dimensions[get_column_letter(col)].width = 18
    ws1.column_dimensions['C'].width = 40  # 主页简介列加宽

    # 工作表2：评论明细
    ws2 = wb.create_sheet("评论明细", 1)
    headers2 = ["用户ID", "昵称", "评论内容", "笔记标题", "帖子链接", "评论时间", "关键词"]
    for col, h in enumerate(headers2, 1):
        ws2.cell(row=1, column=col, value=h)
        ws2.cell(row=1, column=col).font = Font(bold=True)
    row = 2
    for u in users:
        uid, nickname = u.get("user_id", ""), u.get("nickname", "")
        for c in (u.get("comments") or []):
            note_id = c.get("note_id", "")
            token = c.get("note_xsec_token", "") or ""
            source = c.get("note_xsec_source", "") or ""
            note_url = f"https://www.xiaohongshu.com/explore/{note_id}" if note_id else ""
            if token:
                note_url = f"{note_url}?xsec_token={token}&xsec_source={source}" if note_url else ""
            ws2.cell(row=row, column=1, value=uid)
            ws2.cell(row=row, column=2, value=nickname)
            ws2.cell(row=row, column=3, value=c.get("content", ""))
            ws2.cell(row=row, column=4, value=c.get("note_title", ""))
            ws2.cell(row=row, column=5, value=note_url)
            ws2.cell(row=row, column=6, value=c.get("comment_time_str", "") or "")
            ws2.cell(row=row, column=7, value=c.get("keyword", ""))
            row += 1
    for col in range(1, 8):
        ws2.column_dimensions[get_column_letter(col)].width = 22

    from io import BytesIO
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    filename = f"用户数据_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return send_file(
        buffer,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=filename,
    )


def _link_browser_worker():
    """在独立线程中运行 Playwright，用于打开用户/帖子链接"""
    from playwright.sync_api import sync_playwright
    try:
        with sync_playwright() as p:
            browser = p.chromium.launch(headless=False)
            page = browser.new_page()
            page.set_viewport_size({"width": 1200, "height": 800})
            _link_browser_ready.set()
            while True:
                try:
                    url = _link_browser_url_queue.get(timeout=0.5)
                except thread_queue.Empty:
                    continue
                try:
                    page.goto(url, wait_until="domcontentloaded", timeout=15000)
                except Exception as e:
                    print(f"[链接浏览器] 打开失败 {url}: {e}")
            browser.close()
    except Exception as e:
        print(f"[链接浏览器] 启动失败: {e}")
    finally:
        _link_browser_ready.clear()


@app.route('/api/open_in_browser', methods=['POST'])
def open_in_browser():
    """在 Playwright 浏览器中打开指定 URL（用户主页或帖子链接）"""
    global _link_browser_thread
    data = request.json or {}
    url = (data.get('url') or '').strip()
    if not url or not url.startswith('http'):
        return jsonify({'success': False, 'error': '无效的 URL'}), 400
    if 'xiaohongshu.com' not in url:
        return jsonify({'success': False, 'error': '仅支持小红书链接'}), 400
    if _link_browser_thread is None or not _link_browser_thread.is_alive():
        _link_browser_ready.clear()
        _link_browser_thread = threading.Thread(target=_link_browser_worker, daemon=True)
        _link_browser_thread.start()
        if not _link_browser_ready.wait(timeout=15):
            return jsonify({'success': False, 'error': 'Playwright 浏览器启动超时'}), 500
    try:
        _link_browser_url_queue.put_nowait(url)
        return jsonify({'success': True, 'message': '已在浏览器中打开'})
    except thread_queue.Full:
        return jsonify({'success': False, 'error': '请求过多'}), 503


def run_crawler(keywords, max_notes=100, max_comments=100, comment_filter_keywords=[]):
    """运行爬虫（两级抓取）"""
    global crawler_status
    
    crawler_status['running'] = True
    crawler_status['total_users'] = 0
    crawler_status['progress'] = 0
    
    try:
        crawler = XHSCrawler()
        crawler.init_browser()  # 初始化浏览器
        total_keywords = len(keywords)
        
        for idx, keyword in enumerate(keywords):
            crawler_status['current_keyword'] = keyword
            crawler_status['progress'] = int((idx / total_keywords) * 100)
            crawler_status['message'] = f'正在爬取关键词: {keyword} (帖子数: {max_notes}, 评论数: {max_comments})'
            
            # 第一级：搜索关键词，获取最新帖子
            notes = crawler.search_notes_by_keyword(keyword, max_notes=max_notes)
            print(f"[爬虫] 关键词「{keyword}」 共获取 {len(notes)} 个帖子")
            if notes:
                print("[爬虫] 帖子列表原始数据示例(第一条):", json.dumps(notes[0], ensure_ascii=False, indent=2))
            
            crawler_status['message'] = f'找到 {len(notes)} 个帖子，开始抓取评论...'
            
            # 第二级：对每个帖子抓取评论
            for note_idx, note in enumerate(notes):
                note_progress = int((note_idx / len(notes)) * 100) if notes else 0
                crawler_status['message'] = f'正在处理帖子 {note_idx + 1}/{len(notes)}: {note.get("note_id", "")[:8]}...'
                
                # 抓取评论
                comments_data = crawler.crawl_note_comments(
                    note.get('note_id'),
                    max_comments=max_comments,
                    filter_keywords=comment_filter_keywords,
                    xsec_token=note.get('xsec_token', ''),
                    xsec_source=note.get('xsec_source', 'pc_search')
                )
                
                # 控制台日志：本帖子评论条数 + 单条评论原始数据示例
                print(f"[爬虫] 帖子 {note.get('note_id', '')[:12]}... 评论数: {len(comments_data)}")
                if comments_data:
                    print("[爬虫] 单条评论原始数据示例:", json.dumps(comments_data[0], ensure_ascii=False, indent=2))
                
                # 保存评论和用户信息
                for comment_data in comments_data:
                    user = comment_data.get('user', {})
                    if user and user.get('user_id'):
                        user_id = user['user_id']
                        user_file = os.path.join(USERS_DIR, f"{user_id}.json")
                        is_new_user = not os.path.exists(user_file)
                        
                        # 控制台日志：用户原始信息（新用户时打完整，老用户只打一条简短日志）
                        if is_new_user:
                            print(f"[爬虫] 新用户原始信息 user_id={user_id}:", json.dumps(user, ensure_ascii=False, indent=2))
                        else:
                            print(f"[爬虫] 已有用户追加评论 user_id={user_id} nickname={user.get('nickname', '')}")
                        
                        # 检查用户是否已存在，如果存在则更新评论列表
                        if os.path.exists(user_file):
                            with open(user_file, 'r', encoding='utf-8') as f:
                                existing_user = json.load(f)
                        else:
                            existing_user = user.copy()
                        
                        # 添加评论信息
                        if 'comments' not in existing_user:
                            existing_user['comments'] = []
                        
                        # 评论发布时间：小红书 API 返回的 time 为毫秒时间戳（如 1771346050000）
                        ts = comment_data.get('time') or 0
                        if ts:
                            sec = int(ts) / 1000 if int(ts) >= 1e12 else int(ts)
                            comment_time_str = datetime.fromtimestamp(sec).strftime('%Y-%m-%d %H:%M:%S')
                        else:
                            comment_time_str = ''
                        
                        existing_user['comments'].append({
                            'comment_id': comment_data.get('comment_id'),
                            'content': comment_data.get('content'),
                            'note_id': note.get('note_id'),
                            'note_title': note.get('title'),
                            'note_xsec_token': note.get('xsec_token', ''),
                            'note_xsec_source': note.get('xsec_source', ''),
                            'keyword': keyword,
                            'comment_time': ts,
                            'comment_time_str': comment_time_str,
                            'crawl_time': datetime.now().isoformat()
                        })
                        
                        existing_user['keyword'] = keyword
                        existing_user['crawl_time'] = datetime.now().isoformat()
                        # 用户主页需 /user/profile/ 且带 xsec_token、xsec_source 才能正常跳转，用当前帖子参数
                        xsec_source = note.get('xsec_source') or 'pc_note'
                        xsec_token = note.get('xsec_token') or ''
                        qs = f"xsec_source={xsec_source}"
                        if xsec_token:
                            qs = f"xsec_token={xsec_token}&{qs}"
                        existing_user['user_url'] = f"https://www.xiaohongshu.com/user/profile/{user_id}?{qs}"
                        
                        # 若尚未有主页简介，则请求用户主页抓取简介
                        if not (existing_user.get('desc') or existing_user.get('user_desc')):
                            try:
                                creator_info = crawler.get_creator_info(user_id, xsec_token, xsec_source)
                                desc = XHSCrawler.get_creator_desc(creator_info)
                                if desc:
                                    existing_user['desc'] = desc
                                time.sleep(1)  # 避免请求过快
                            except Exception as e:
                                print(f"[爬虫] 获取用户简介失败 user_id={user_id}: {e}")
                        
                        with open(user_file, 'w', encoding='utf-8') as f:
                            json.dump(existing_user, f, ensure_ascii=False, indent=2)
                        
                        crawler_status['total_users'] += 1
                
                # 更新进度
                keyword_progress = int((idx / total_keywords) * 100)
                note_progress = int((note_idx / len(notes)) * 100) if notes else 0
                crawler_status['progress'] = keyword_progress + int(note_progress / total_keywords)
            
            crawler_status['progress'] = int(((idx + 1) / total_keywords) * 100)
        
        crawler_status['message'] = f'爬取完成！共获取 {crawler_status["total_users"]} 个用户'
        crawler_status['current_keyword'] = ''
        
    except Exception as e:
        error_msg = str(e)
        if "Playwright 浏览器未安装" in error_msg or "Executable doesn't exist" in error_msg:
            crawler_status['message'] = '错误: Playwright 浏览器未安装。请运行 "playwright install chromium" 或 "安装依赖.bat"'
        elif "等待登录超时" in error_msg:
            crawler_status['message'] = '未在限定时间内完成登录。请先在小红书页面扫码/登录，再点击「开始爬取」。'
        else:
            crawler_status['message'] = f'爬取出错: {error_msg}'
        import traceback
        traceback.print_exc()
    finally:
        if 'crawler' in locals():
            try:
                crawler.close()  # 关闭浏览器
            except:
                pass
        crawler_status['running'] = False
        crawler_status['progress'] = 100

if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0', port=5000)
